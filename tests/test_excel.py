# tests/test_excel.py
import io
import pytest
import openpyxl
from extractors.excel import extract_excel_facts


def _make_placement_slip() -> bytes:
    wb = openpyxl.Workbook()

    # Policy Info sheet
    ws1 = wb.active
    ws1.title = "Policy Info"
    ws1.append(["Modular Policy Placement Slip Details", "Value", "Notes", "x", "Rules"])
    ws1.append(["Corporate Name", "ACME Corp", None, "y", None])
    ws1.append(["Insurer Name", "Star Health", None, "y", None])
    ws1.append(["Net Premium", "5000000", None, "y", None])

    # Modular Plan sheet
    ws2 = wb.create_sheet("Modular Plan")
    ws2.append([None, "Plan number", "-", "Integer", None, 1, 2])
    ws2.append([None, "Plan names", "-", "Text", None, "Silver", "Gold"])
    ws2.append([None, None, None, None, None, None, None])
    ws2.append([None, None, None, None, None, None, None])
    ws2.append([None, None, None, None, None, None, None])
    ws2.append(["Priority", "Sub-benefits category", "Sub-benefits", None, "Flag", None, None])
    ws2.append([1, "General Coverage", "Base SI", "Integer", "Yes", "500000", "800000"])
    ws2.append([2, "General Coverage", "Room Rent Restriction", "Text", "Yes", "1% of SI", "Single Pvt AC Room"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_rater() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raters"
    ws.append(["Benefit ID", "Benefit name (for reference only)", "Active", "Pro-Rated", "Rater Type",
               "Property 1", "Operator 1", "Value 1", "Property 2", "Operator 2", "Value 2",
               "Property 3", "Operator 3", "Value 3", "Delimiter", "Premium",
               "Employee Premium", "Employee Premium Pro-Rated", "HR Premium", "HR Premium Pro-Rated",
               "Insurer Premium", "Insurer Premium Pro-Rated", "Additional Fixed Cost", "Tax Rate"])
    ws.append(["BEN-001", "Dental Care Plan", True, "No", "life",
               None, None, None, None, None, None, None, None, None, ";", 470.0,
               398.0, "No", 398.0, "No", 398.0, "No", 0, 18])
    ws.append(["BEN-002", "Vision Care Plan", True, "No", "life",
               None, None, None, None, None, None, None, None, None, ";", 199.0,
               168.0, "No", 168.0, "No", 168.0, "No", 0, 18])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_extract_policy_info():
    facts = extract_excel_facts(_make_placement_slip())
    assert "policy_info" in facts
    assert facts["policy_info"]["Corporate Name"] == "ACME Corp"
    assert facts["policy_info"]["Insurer Name"] == "Star Health"


def test_extract_plans():
    facts = extract_excel_facts(_make_placement_slip())
    assert "plans" in facts
    assert "Silver" in facts["plans"]
    assert "Gold" in facts["plans"]
    assert facts["plans"]["Silver"]["Base SI"] == "500000"
    assert facts["plans"]["Gold"]["Room Rent Restriction"] == "Single Pvt AC Room"


def test_extract_rater_benefits():
    facts = extract_excel_facts(_make_rater())
    assert "benefits" in facts
    names = [b["name"] for b in facts["benefits"]]
    assert "Dental Care Plan" in names
    dental = next(b for b in facts["benefits"] if b["name"] == "Dental Care Plan")
    assert dental["premium"] == 470.0


def test_unknown_sheet_returns_empty():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Random"
    ws.append(["foo", "bar"])
    buf = io.BytesIO()
    wb.save(buf)
    facts = extract_excel_facts(buf.getvalue())
    assert facts == {}
