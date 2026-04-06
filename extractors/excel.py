import io
import openpyxl


def _cell_str(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s in ("None", "nan") else s


def extract_excel_facts(file_bytes: bytes) -> dict:
    """Extract all readable text from every sheet. Sheet-name agnostic."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        return {}

    sheets_text = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []
        for row in ws.iter_rows(values_only=True):
            cells = [_cell_str(c) for c in row if _cell_str(c)]
            if cells:
                rows_text.append(" | ".join(cells))
        if rows_text:
            sheets_text[sheet_name] = "\n".join(rows_text)

    return {"raw_sheets": sheets_text} if sheets_text else {}
